"""
Generate a Mac-compatible .xlsx template for bulk Raw Materials import.
Client fills the rows; we import into Firestore `rawMaterials` afterwards.

Run: python3 scripts/makeRawMaterialsTemplate.py
Output: imports/Grabio-RawMaterials-Template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = os.path.join(os.getcwd(), "imports")
OUT_PATH = os.path.join(OUT_DIR, "Grabio-RawMaterials-Template.xlsx")

# Column definitions: (header, width, required, help text)
COLUMNS = [
    ("Name", 28, True, "Raw material name (e.g. Mozzarella Cheese)"),
    ("Unit", 14, True, "One of: kg, gram, liter, ml, piece, meter"),
    ("Cost Per Unit (USD)", 18, True, "Cost of ONE unit, must be greater than 0"),
    ("Current Stock", 15, False, "Quantity on hand now (number). Default 0"),
    ("Min Threshold", 15, False, "Alert when stock drops below this. Default 10"),
    ("Reorder Point", 15, False, "Reorder when stock reaches this. Default 20"),
    ("Storage Location", 20, False, "Optional (e.g. Fridge 1, Dry Store)"),
    ("Expiry Tracking (yes/no)", 22, False, "Type yes to track expiry, else leave blank"),
    ("Expiry Date (YYYY-MM-DD)", 24, False, "Optional, only if Expiry Tracking = yes"),
]

UNITS = ["kg", "gram", "liter", "ml", "piece", "meter"]

TITLE_FILL = PatternFill("solid", fgColor="14532D")
HEADER_FILL = PatternFill("solid", fgColor="16A34A")
REQ_FILL = PatternFill("solid", fgColor="DCFCE7")
HELP_FILL = PatternFill("solid", fgColor="F1F5F9")
WHITE = Font(color="FFFFFF", bold=True)
HELP_FONT = Font(color="475569", italic=True, size=9)
thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Materials"

    ncols = len(COLUMNS)

    # Row 1: title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value="GRABIO — RAW MATERIALS IMPORT")
    tcell.font = Font(color="FFFFFF", bold=True, size=14)
    tcell.fill = TITLE_FILL
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: subtitle / instruction
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    scell = ws.cell(
        row=2, column=1,
        value="Fill one row per raw material. Green columns are REQUIRED. Delete the grey EXAMPLE row before sending back.",
    )
    scell.font = HELP_FONT
    scell.fill = HELP_FILL
    scell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    header_row = 3
    help_row = 4
    example_row = 5

    # Header row
    for idx, (name, width, required, _help) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=header_row, column=idx, value=name)
        c.font = WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        col_letter = c.column_letter
        ws.column_dimensions[col_letter].width = width

    # Help row
    for idx, (_name, _width, required, help_text) in enumerate(COLUMNS, start=1):
        label = ("REQUIRED — " if required else "Optional — ") + help_text
        c = ws.cell(row=help_row, column=idx, value=label)
        c.font = HELP_FONT
        c.fill = REQ_FILL if required else HELP_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[help_row].height = 40

    # Example row (grey, to be deleted)
    example = ["Mozzarella Cheese", "kg", 6.50, 20, 10, 20, "Fridge 1", "yes", "2026-12-31"]
    for idx, val in enumerate(example, start=1):
        c = ws.cell(row=example_row, column=idx, value=val)
        c.fill = HELP_FILL
        c.font = Font(color="94A3B8", italic=True)
        c.border = BORDER

    # Unit dropdown validation for a big range (rows 5..500)
    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(UNITS),
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Pick a valid unit: " + ", ".join(UNITS)
    dv.errorTitle = "Invalid unit"
    dv.prompt = "Choose: " + ", ".join(UNITS)
    dv.promptTitle = "Unit"
    ws.add_data_validation(dv)
    dv.add("B%d:B500" % example_row)

    # Freeze header so it stays visible while scrolling
    ws.freeze_panes = "A%d" % example_row

    # Second sheet: instructions
    info = wb.create_sheet("Instructions")
    info.column_dimensions["A"].width = 100
    lines = [
        ("GRABIO — Raw Materials Import Instructions", True),
        ("", False),
        ("1. Go to the 'Raw Materials' tab.", False),
        ("2. Fill ONE row per material, starting from the first empty row.", False),
        ("3. REQUIRED columns (green): Name, Unit, Cost Per Unit.", False),
        ("4. Unit must be one of: kg, gram, liter, ml, piece, meter (use the dropdown).", False),
        ("5. Cost Per Unit must be a number greater than 0 (no currency symbol).", False),
        ("6. Numbers only for Current Stock / Min Threshold / Reorder Point (leave blank for defaults).", False),
        ("7. Expiry Date format: YYYY-MM-DD (example: 2026-12-31). Only if Expiry Tracking = yes.", False),
        ("8. DELETE the grey EXAMPLE row before sending the file back.", False),
        ("9. Do NOT rename the column headers or the sheet tabs.", False),
        ("10. SKU and barcode are generated automatically — do not add them.", False),
        ("", False),
        ("When done, save and send this .xlsx file back.", False),
    ]
    for r, (text, bold) in enumerate(lines, start=1):
        c = info.cell(row=r, column=1, value=text)
        c.font = Font(bold=bold, size=13 if bold else 11)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb.save(OUT_PATH)
    print("Wrote:", OUT_PATH)


if __name__ == "__main__":
    build()
